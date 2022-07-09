import openpyxl
from openpyxl.styles import PatternFill

old_file = openpyxl.load_workbook("old_file.xlsx", keep_vba=True)
new_file = openpyxl.load_workbook("new_file.xlsx", keep_vba=True)
marker_color = PatternFill(fgColor='00008080', fill_type='solid')

for sheetname_old_file, sheetname_new_file in zip(old_file.sheetnames, new_file.sheetnames):
    sheet_old_file = old_file[sheetname_old_file]
    sheet_new_file = new_file[sheetname_new_file]
    # last filled row
    max_row = max(sheet_old_file.max_row, sheet_new_file.max_row)
    # last filled column
    max_column = max(sheet_old_file.max_column, sheet_new_file.max_column)
    print(sheet_old_file, sheet_new_file)
    for col_idx in range(1, max_column + 1):
        for row_idx in range(1, max_row + 1):
            old_cell_value = sheet_old_file.cell(column=col_idx, row=row_idx).value
            new_cell_value = sheet_new_file.cell(column=col_idx, row=row_idx).value
            if old_cell_value != new_cell_value:
                sheet_new_file.cell(column=col_idx, row=row_idx).fill = marker_color
                print('\nUnterschied gefunden:')
                print(f'Spalte: {col_idx}, Zeile: {row_idx}')
                print(f'Alter Wert: {old_cell_value}')
                print(f'Neuer Wert: {new_cell_value}')

        print(f'Spalte: {col_idx} fertig.')

new_file.save('result.xlsx')
new_file.close()
